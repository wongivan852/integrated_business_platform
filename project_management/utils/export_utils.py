"""
Export utilities for generating PDF, Excel, and CSV reports
Supports project reports, analytics dashboards, and resource reports
"""

import csv
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference


def generate_csv_response(data, filename, headers=None):
    """
    Generate CSV file response

    Args:
        data: List of dictionaries or list of lists
        filename: Name of the file (without .csv extension)
        headers: Optional list of headers (if data is list of dicts, uses keys)

    Returns:
        HttpResponse with CSV content
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    writer = csv.writer(response)

    # Write headers
    if headers:
        writer.writerow(headers)
    elif data and isinstance(data[0], dict):
        writer.writerow(data[0].keys())

    # Write data
    for row in data:
        if isinstance(row, dict):
            writer.writerow(row.values())
        else:
            writer.writerow(row)

    return response


def generate_excel_workbook(filename):
    """
    Create a new Excel workbook with basic styling

    Args:
        filename: Name of the file (without .xlsx extension)

    Returns:
        Tuple of (HttpResponse, Workbook)
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    wb = Workbook()
    return response, wb


def style_excel_header(ws, header_row, num_columns):
    """
    Apply styling to Excel header row

    Args:
        ws: Worksheet object
        header_row: Row number of header (1-indexed)
        num_columns: Number of columns to style
    """
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, num_columns + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def add_excel_borders(ws, start_row, end_row, num_columns):
    """
    Add borders to Excel cells

    Args:
        ws: Worksheet object
        start_row: Starting row (1-indexed)
        end_row: Ending row (1-indexed)
        num_columns: Number of columns
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(start_row, end_row + 1):
        for col in range(1, num_columns + 1):
            ws.cell(row=row, column=col).border = thin_border


def export_project_list_excel(projects):
    """
    Export project list to Excel with formatting

    Args:
        projects: QuerySet of Project objects

    Returns:
        HttpResponse with Excel file
    """
    response, wb = generate_excel_workbook(f'projects_{datetime.now().strftime("%Y%m%d")}')
    ws = wb.active
    ws.title = "Projects"

    # Headers
    headers = [
        'Project Code', 'Name', 'Status', 'Priority', 'Owner',
        'Start Date', 'End Date', 'Progress %', 'Budget', 'Actual Cost',
        'Team Size', 'Tasks Total', 'Tasks Completed'
    ]
    ws.append(headers)
    style_excel_header(ws, 1, len(headers))

    # Data rows
    for project in projects:
        ws.append([
            project.project_code,
            project.name,
            project.get_status_display(),
            project.get_priority_display(),
            project.owner.get_full_name() or project.owner.email,
            project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
            project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
            f"{project.progress_percentage}%",
            float(project.budget) if project.budget else 0,
            float(project.actual_cost),
            project.team_members.count(),
            project.tasks.count(),
            project.tasks.filter(status='completed').count()
        ])

    # Styling
    add_excel_borders(ws, 1, ws.max_row, len(headers))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response


def export_project_analytics_pdf(project, metrics):
    """
    Export project analytics to PDF report

    Args:
        project: Project object
        metrics: Dictionary of analytics metrics

    Returns:
        HttpResponse with PDF file
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="project_analytics_{project.project_code}_{datetime.now().strftime("%Y%m%d")}.pdf"'

    # Create PDF document
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12
    )

    # Build content
    content = []

    # Title
    content.append(Paragraph(f"Project Analytics Report", title_style))
    content.append(Paragraph(f"{project.name}", styles['Heading2']))
    content.append(Spacer(1, 0.3*inch))

    # Project Overview
    content.append(Paragraph("Project Overview", heading_style))
    overview_data = [
        ['Project Code:', project.project_code],
        ['Status:', project.get_status_display()],
        ['Priority:', project.get_priority_display()],
        ['Owner:', project.owner.get_full_name() or project.owner.email],
        ['Start Date:', project.start_date.strftime('%Y-%m-%d') if project.start_date else 'N/A'],
        ['End Date:', project.end_date.strftime('%Y-%m-%d') if project.end_date else 'N/A'],
        ['Progress:', f"{project.progress_percentage}%"]
    ]
    overview_table = Table(overview_data, colWidths=[2*inch, 4*inch])
    overview_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0'))
    ]))
    content.append(overview_table)
    content.append(Spacer(1, 0.3*inch))

    # Metrics
    content.append(Paragraph("Key Metrics", heading_style))
    metrics_data = [
        ['Metric', 'Value'],
        ['Total Tasks', str(metrics.get('total_tasks', 0))],
        ['Completed Tasks', str(metrics.get('completed_tasks', 0))],
        ['In Progress Tasks', str(metrics.get('in_progress_tasks', 0))],
        ['Overdue Tasks', str(metrics.get('overdue_tasks', 0))],
        ['Completion Rate', f"{metrics.get('completion_rate', 0)}%"],
        ['Budget', f"${metrics.get('budget', 0):,.2f}"],
        ['Actual Cost', f"${metrics.get('actual_cost', 0):,.2f}"],
        ['Cost Variance', f"${metrics.get('cost_variance', 0):,.2f}"],
        ['Health Score', f"{metrics.get('health_score', 0)}/100"]
    ]
    metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
    metrics_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
    ]))
    content.append(metrics_table)
    content.append(Spacer(1, 0.3*inch))

    # Team Members
    if project.team_members.exists():
        content.append(Paragraph("Team Members", heading_style))
        team_data = [['Name', 'Role', 'Email']]
        for member in project.projectmember_set.all()[:10]:  # Limit to 10
            team_data.append([
                member.user.get_full_name() or member.user.email,
                member.get_role_display(),
                member.user.email
            ])
        team_table = Table(team_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
        team_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke)
        ]))
        content.append(team_table)

    # Footer
    content.append(Spacer(1, 0.5*inch))
    content.append(Paragraph(
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        styles['Normal']
    ))

    # Build PDF
    doc.build(content)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


def export_tasks_excel(tasks, project_name):
    """
    Export task list to Excel

    Args:
        tasks: QuerySet of Task objects
        project_name: Name of project (for filename)

    Returns:
        HttpResponse with Excel file
    """
    response, wb = generate_excel_workbook(
        f'tasks_{project_name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}'
    )
    ws = wb.active
    ws.title = "Tasks"

    # Headers
    headers = [
        'Task Code', 'Title', 'Status', 'Priority', 'Progress %',
        'Start Date', 'End Date', 'Due Date', 'Assigned To',
        'Estimated Hours', 'Actual Hours', 'Is Milestone'
    ]
    ws.append(headers)
    style_excel_header(ws, 1, len(headers))

    # Data rows
    for task in tasks:
        assigned_to = ', '.join([u.get_full_name() or u.email for u in task.assigned_to.all()])
        ws.append([
            task.task_code,
            task.title,
            task.get_status_display(),
            task.get_priority_display(),
            f"{task.progress}%",
            task.start_date.strftime('%Y-%m-%d') if task.start_date else '',
            task.end_date.strftime('%Y-%m-%d') if task.end_date else '',
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '',
            assigned_to,
            float(task.estimated_hours) if task.estimated_hours else 0,
            float(task.actual_hours) if task.actual_hours else 0,
            'Yes' if task.is_milestone else 'No'
        ])

    # Styling
    add_excel_borders(ws, 1, ws.max_row, len(headers))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response


def export_resource_allocation_excel(resources, start_date, end_date):
    """
    Export resource allocation report to Excel

    Args:
        resources: QuerySet of Resource objects with assignments
        start_date: Start date for report
        end_date: End date for report

    Returns:
        HttpResponse with Excel file
    """
    response, wb = generate_excel_workbook(
        f'resource_allocation_{datetime.now().strftime("%Y%m%d")}'
    )
    ws = wb.active
    ws.title = "Resource Allocation"

    # Headers
    headers = [
        'Resource Name', 'Job Title', 'Department', 'Hourly Rate',
        'Availability %', 'Available Hours/Day', 'Assigned Projects',
        'Total Assigned Hours', 'Total Actual Hours', 'Utilization %'
    ]
    ws.append(headers)
    style_excel_header(ws, 1, len(headers))

    # Data rows
    for resource in resources:
        assignments = resource.resourceassignment_set.filter(
            start_date__lte=end_date,
            end_date__gte=start_date
        )

        total_assigned = sum([a.assigned_hours for a in assignments])
        total_actual = sum([a.actual_hours for a in assignments])
        projects = ', '.join(set([a.task.project.name for a in assignments]))

        utilization = 0
        if resource.available_hours_per_day > 0:
            # Calculate utilization based on working days in period
            days = (end_date - start_date).days + 1
            available_hours = resource.available_hours_per_day * days
            utilization = (total_assigned / available_hours * 100) if available_hours > 0 else 0

        ws.append([
            resource.user.get_full_name() or resource.user.email,
            resource.job_title,
            resource.department,
            float(resource.hourly_rate),
            float(resource.availability_percentage),
            float(resource.available_hours_per_day),
            projects,
            float(total_assigned),
            float(total_actual),
            f"{utilization:.1f}%"
        ])

    # Styling
    add_excel_borders(ws, 1, ws.max_row, len(headers))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response


def export_portfolio_analytics_excel(projects_data):
    """
    Export portfolio analytics to Excel with multiple sheets

    Args:
        projects_data: List of dictionaries with project analytics data

    Returns:
        HttpResponse with Excel file
    """
    response, wb = generate_excel_workbook(
        f'portfolio_analytics_{datetime.now().strftime("%Y%m%d")}'
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Portfolio Summary"

    # Calculate summary metrics
    total_projects = len(projects_data)
    total_budget = sum([p.get('budget', 0) for p in projects_data])
    total_cost = sum([p.get('actual_cost', 0) for p in projects_data])
    avg_health = sum([p.get('health_score', 0) for p in projects_data]) / total_projects if total_projects > 0 else 0

    summary_data = [
        ['Portfolio Analytics Summary', ''],
        ['', ''],
        ['Metric', 'Value'],
        ['Total Projects', total_projects],
        ['Total Budget', f"${total_budget:,.2f}"],
        ['Total Actual Cost', f"${total_cost:,.2f}"],
        ['Cost Variance', f"${total_budget - total_cost:,.2f}"],
        ['Average Health Score', f"{avg_health:.1f}/100"],
        ['', ''],
        ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]

    for row in summary_data:
        ws_summary.append(row)

    # Style summary
    ws_summary['A1'].font = Font(bold=True, size=16, color="366092")
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    ws_summary.merge_cells('A1:B1')

    style_excel_header(ws_summary, 3, 2)

    # Sheet 2: Projects Detail
    ws_detail = wb.create_sheet("Projects Detail")
    headers = [
        'Project Code', 'Name', 'Status', 'Health Score',
        'Progress %', 'Budget', 'Actual Cost', 'Cost Variance',
        'Tasks Total', 'Tasks Completed', 'Overdue Tasks'
    ]
    ws_detail.append(headers)
    style_excel_header(ws_detail, 1, len(headers))

    for project in projects_data:
        ws_detail.append([
            project.get('project_code', ''),
            project.get('name', ''),
            project.get('status', ''),
            project.get('health_score', 0),
            f"{project.get('progress', 0)}%",
            float(project.get('budget', 0)),
            float(project.get('actual_cost', 0)),
            float(project.get('budget', 0)) - float(project.get('actual_cost', 0)),
            project.get('tasks_total', 0),
            project.get('tasks_completed', 0),
            project.get('tasks_overdue', 0)
        ])

    add_excel_borders(ws_detail, 1, ws_detail.max_row, len(headers))

    # Auto-adjust column widths for both sheets
    for ws in [ws_summary, ws_detail]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response
